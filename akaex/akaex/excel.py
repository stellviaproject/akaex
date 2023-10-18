import openpyxl
from django.db.models import fields
from openpyxl import Workbook
from io import BytesIO

def queryset_to_excel_bytes(queryset):
    wb = Workbook()
    ws = wb.active

    # Write the headers
    for col_num, field in enumerate(queryset.model._meta.fields):
        col_letter = openpyxl.utils.get_column_letter(col_num + 1)
        ws[col_letter + '1'] = field.name

    # Write the data
    for row_num, obj in enumerate(queryset, start=2):
        for col_num, field in enumerate(queryset.model._meta.fields):
            col_letter = openpyxl.utils.get_column_letter(col_num + 1)
            ws[col_letter + str(row_num)] = getattr(obj, field.name)

    # Save the workbook to a BytesIO object
    output = BytesIO()
    wb.save(output)

    # Reset the output BytesIO object to the beginning
    output.seek(0)

    return output.read()

import openpyxl
from io import BytesIO

def excel_bytes_to_queryset(excel_bytes, ModelClass, unique_fields):
    wb = openpyxl.load_workbook(BytesIO(excel_bytes))
    ws = wb.active

    # Get the headers
    headers = []
    for col_num in range(1, ws.max_column + 1):
        headers.append(ws.cell(row=1, column=col_num).value)

    # Get the data
    for row_num in range(2, ws.max_row + 1):
        row_data = {}
        for col_num, header in enumerate(headers, start=1):
            row_data[header] = ws.cell(row=row_num, column=col_num).value

        # Update or create the object
        object, created = ModelClass.objects.update_or_create(
            defaults=row_data,
            **{field: row_data[field] for field in unique_fields}
        )
        object.save()
